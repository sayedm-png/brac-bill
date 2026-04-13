import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime
import re
from io import BytesIO

# --- 1. CONFIGURATION ---
# The public link you provided for distance allowances
DRIVE_URL = "https://drive.google.com/file/d/1yUpT0N-8d_P1org-e5LwfffCYxco2sPb/view?ts=69dc8c08"
TEMPLATE_FILE = "Template.xlsm"

# BRAC Official Rates
MEAL_RATES = {"B": 70, "L": 140, "D": 140}
HALT_RATE = 150

# --- 2. DATA FETCHING ---
@st.cache_data(ttl=600)
def get_allowances(url):
    try:
        file_id = re.search(r'[-\w]{25,}', url).group()
        csv_link = f'https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv'
        df = pd.read_csv(csv_link)
        df.columns = [c.strip() for c in df.columns]
        return df
    except Exception as e:
        st.sidebar.error(f"Sync Error: Ensure Google Drive file is set to 'Anyone with the link can view'")
        return None

# --- 3. APP UI ---
st.set_page_config(page_title="BRAC Bill Pro", layout="wide")
st.title("💸 Frontier Tech Billing App")

allowance_df = get_allowances(DRIVE_URL)

with st.sidebar:
    st.header("Trip Input")
    visit_date = st.date_input("Date of Visit", datetime.date.today())
    
    if allowance_df is not None:
        area = st.selectbox("Area of Visit", options=allowance_df['Area'].unique())
        fixed_dist = allowance_df.loc[allowance_df['Area'] == area, 'Allowance'].values[0]
        st.info(f"Fixed Distance for {area}: {fixed_dist} BDT")
    else:
        area = st.text_input("Enter Area Manually")
        fixed_dist = st.number_input("Fixed Distance", 0)

    ground_travel = st.number_input("Ground Travel (Actual)", 0)
    
    st.divider()
    st.subheader("Meals & Stay")
    b = st.number_input("Breakfast (70)", 0)
    l = st.number_input("Lunch (140)", 0)
    d = st.number_input("Dinner (140)", 0)
    halt_days = st.number_input("Night Haltage (150)", 0)

# --- 4. CALCULATIONS ---
food_total = (b * MEAL_RATES["B"]) + (l * MEAL_RATES["L"]) + (d * MEAL_RATES["D"])
halt_total = halt_days * HALT_RATE
grand_total = ground_travel + fixed_dist + food_total + halt_total

# --- 5. PREVIEW & EXPORT ---
col_pre, col_res = st.columns([2, 1])

with col_pre:
    st.subheader("Billing Summary")
    summary = pd.DataFrame({
        "Category": ["Travel (Ground)", "Distance Allowance", "Meals Total", "Night Haltage"],
        "Amount (BDT)": [ground_travel, fixed_dist, food_total, halt_total]
    })
    st.table(summary)
    st.metric("Grand Total", f"{grand_total} BDT")

with col_res:
    st.subheader("Generate File")
    if st.button("Apply to Template"):
        try:
            # Load template from GitHub folder
            wb = load_workbook(TEMPLATE_FILE, keep_vba=True)
            ws = wb.active

            # --- CELL UPDATES (Update these to your exact Excel cells) ---
            ws['B5'] = str(visit_date)  # Example cell for Date
            ws['B6'] = area             # Example cell for Area
            ws['E10'] = ground_travel   # Example cell for Ground Travel
            ws['E11'] = fixed_dist      # Example cell for Distance Allowance
            ws['E12'] = food_total      # Example cell for Meals
            ws['E13'] = halt_total      # Example cell for Haltage

            # Save to memory
            output = BytesIO()
            wb.save(output)
            
            st.success("Format populated!")
            st.download_button(
                label="📥 Download Official .xlsm",
                data=output.getvalue(),
                file_name=f"Bill_{area}_{visit_date}.xlsm",
                mime="application/vnd.ms-excel.sheet.macroenabled.12"
            )
        except FileNotFoundError:
            st.error("Error: 'Template.xlsm' not found in GitHub. Please upload it.")
        except Exception as e:
            st.error(f"Logic Error: {e}")
